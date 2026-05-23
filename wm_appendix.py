#!/usr/bin/env python3
"""
wm_appendix.py - WM Appendix PDF Landscape + XLSX v3.0
Input:  /home/claude/market_data.json
Output: /home/claude/financial_wm_appendix_{DATE_STR}.pdf
        /home/claude/financial_daily_data_{DATE_STR}.xlsx
"""
import json, os
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WHITE=HexColor("#FFFFFF"); NAVY=HexColor("#0B2545"); GOLD=HexColor("#C9A84C")
AMBER=HexColor("#E07B00"); GREEN=HexColor("#1A8A4A"); RED=HexColor("#C0392B")
LGREY=HexColor("#F5F7FA"); GREY=HexColor("#2C3E50")
NAVY_H="0B2545"; GOLD_H="C9A84C"; GREEN_H="1A8A4A"; RED_H="C0392B"; AMBER_H="E07B00"

with open("/home/claude/market_data.json","r",encoding="utf-8") as f: mkt=json.load(f)
meta=mkt.get("meta",{}); forex=mkt.get("forex",{}); idx=mkt.get("indices",{})
bonds=mkt.get("bonds",{}); comm=mkt.get("commodities",{}); vol=mkt.get("volatility",{})
macro=mkt.get("macro",{}); crypto=mkt.get("crypto",{})
DATE_REPORT=meta.get("DATE_REPORT","N/D"); DATE_CHIUSURA=meta.get("DATE_CHIUSURA","N/D"); DATE_STR=meta.get("DATE_STR","2026-01-01")
PDF_PATH=f"/home/claude/financial_wm_appendix_{DATE_STR}.pdf"; PDF_OUT=f"/mnt/user-data/outputs/financial_wm_appendix_{DATE_STR}.pdf"
XLSX_PATH=f"/home/claude/financial_daily_data_{DATE_STR}.xlsx"; XLSX_OUT=f"/mnt/user-data/outputs/financial_daily_data_{DATE_STR}.xlsx"

def ps(n,**kw): return ParagraphStyle(n,**kw)
SB=ps("b",fontName="Helvetica",fontSize=8,textColor=GREY,leading=11)
SW=ps("w",fontName="Helvetica-Bold",fontSize=8,textColor=WHITE,leading=11)
SS=ps("s",fontName="Helvetica",fontSize=7,textColor=GREY,leading=10)
ST=ps("t",fontName="Helvetica-Bold",fontSize=10,textColor=NAVY,leading=13)
SG=ps("g",fontName="Helvetica-Bold",fontSize=8,textColor=GREEN,leading=11)
SR=ps("r",fontName="Helvetica-Bold",fontSize=8,textColor=RED,leading=11)
SA=ps("a",fontName="Helvetica-Bold",fontSize=8,textColor=AMBER,leading=11)

def v(d,k,fmt=None):
    val=d.get(k)
    if val is None or val=="N/D": return "N/D"
    if fmt:
        try: return fmt.format(float(val))
        except: return str(val)
    return str(val)

def hdr_ftr(canvas,doc):
    w,h=landscape(A4); canvas.saveState()
    canvas.setFillColor(NAVY); canvas.rect(0,h-1.5*cm,w,1.5*cm,fill=1,stroke=0)
    canvas.setFillColor(WHITE); canvas.setFont("Helvetica-Bold",10)
    canvas.drawString(1.5*cm,h-1.0*cm,"FINANCIAL DAILY REPORT  di Ettore Turchet - Allegato WM")
    canvas.setFont("Helvetica",8); canvas.drawRightString(w-1.5*cm,h-1.0*cm,f"Chiusura: {DATE_CHIUSURA}  |  Report: {DATE_REPORT}")
    canvas.setStrokeColor(GOLD); canvas.setLineWidth(1.5); canvas.line(1.5*cm,h-1.55*cm,w-1.5*cm,h-1.55*cm)
    canvas.setFont("Helvetica",7); canvas.setFillColor(GREY)
    canvas.drawString(1.5*cm,0.7*cm,"Allegato WM - uso interno")
    canvas.drawRightString(w-1.5*cm,0.7*cm,f"Pag. {doc.page}"); canvas.restoreState()

def sec(t,n):
    return [HRFlowable(width="100%",thickness=1.5,color=GOLD,spaceAfter=4),Paragraph(f"<b>{n} - {t}</b>",ST),Spacer(1,4)]

def tbl(rows,widths,hdr=None):
    tr=[]
    if hdr: tr.append([Paragraph(str(c),SW) for c in hdr])
    for row in rows: tr.append([Paragraph(str(c),SB) if isinstance(c,str) else c for c in row])
    sty=[("BACKGROUND",(0,0),(-1,0),NAVY) if hdr else ("BACKGROUND",(0,0),(-1,-1),WHITE),
         ("ROWBACKGROUNDS",(0,1 if hdr else 0),(-1,-1),[WHITE,LGREY]),
         ("GRID",(0,0),(-1,-1),0.3,HexColor("#CCCCCC")),
         ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
         ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),("WORDWRAP",(0,0),(-1,-1),"LTR")]
    return Table(tr,colWidths=widths,style=TableStyle(sty),repeatRows=1 if hdr else 0)

story=[]
story+=sec("MARKET DATA EXTENDED","A")
story.append(tbl([
    ["FTSE MIB",v(idx,"FTSE_MIB"),v(idx,"FTSE_MIB_DELTA"),"Equity EU","MilanoFinanza","T2"],
    ["DAX",v(idx,"DAX"),v(idx,"DAX_DELTA"),"Equity EU","MilanoFinanza","T2"],
    ["S&P 500",v(idx,"SP500"),v(idx,"SP500_DELTA"),"Equity US","CNBC","T2"],
    ["EUR/USD",v(forex,"EUR_USD"),"N/D","FX","ECB XML","T1"],
    ["EUR/JPY",v(forex,"EUR_JPY"),"N/D","FX","ECB XML","T1"],
    ["BTP 10Y",v(bonds,"BTP_10Y","{}%"),"N/D","Bond IT","MilanoFinanza","T2"],
    ["Bund 10Y",v(bonds,"BUND_10Y","{}%"),"N/D","Bond DE","web_search","T2"],
    ["Spread",v(bonds,"SPREAD_BTP_BUND","{}bp"),"N/D","Spread","calcolato","â€”"],
    ["UST 10Y",v(bonds,"UST_10Y","{}%"),"N/D","Bond US","web_search","T2"],
    ["Brent",v(comm,"BRENT"),v(comm,"BRENT_DELTA"),"Commodity","CNBC","T2"],
    ["Gold",v(comm,"GOLD"),v(comm,"GOLD_DELTA"),"Commodity","CNBC","T2"],
    ["VIX",v(vol,"VIX"),"N/D","Volatilita","CBOE","T1"],
    ["BTC/USD",v(crypto,"BTC_USD"),v(crypto,"BTC_DELTA"),"Crypto","Composio","T2"],
],[4*cm,3*cm,2.5*cm,3.5*cm,4*cm,1.5*cm],hdr=["Asset","Valore","Delta%","Categoria","Fonte","Tier"]))
story.append(Spacer(1,10))

story+=sec("RACCOMANDAZIONI WEALTH MANAGEMENT","B")
story.append(tbl([
    ["Equity EU","Neutrale","->","Spread BTP stabile; FTSE MIB resistenza 50k","Geopolitica Hormuz"],
    ["Equity US","Sottopeso","<-","Valutazioni stretched; insider selling Mag7","Fed hawkish"],
    ["Fixed Income EU","Sovrappeso","->","BTP carry positivo; BCE hawkish supporta","Duration risk"],
    ["Fixed Income US","Neutrale","->","UST 30Y >5%: opportunita selettiva","Deficit USA"],
    ["Commodities Oil","Neutrale","->","Hormuz: scenario B dominante","Geopolitica"],
    ["Gold","Sovrappeso","->","Safe haven + de-dollarizzazione strutturale","Liquidita"],
    ["Cash","Neutrale","->","EUR-STR ~2.9%: parcheggio remunerato","â€”"],
],[4*cm,3*cm,1.5*cm,8.5*cm,4.5*cm],hdr=["Asset Class","Posizione","Dir.","Rationale","Rischio"]))
story.append(Spacer(1,10))

story+=sec("STRESS TEST - SCENARI","C")
try:
    bf=float(str(comm.get("BRENT","0")).replace("$","").replace(",","."))
    ta=f"<${bf*0.87:.1f}"; tb=f"${bf*0.95:.1f}-${bf*1.05:.1f}"; tc=">$140"
except: ta="N/D"; tb="N/D"; tc=">$140"
story.append(tbl([
    ["A - De-escalation","40%",ta,"FTSE MIB +3-5%","Negoziato Iran-USA","Basso"],
    ["B - Stallo","45%",tb,"Laterale +-2%","Hormuz pedaggi","Medio"],
    ["C - Escalation","10%",tc,"FTSE MIB -8-10%","Attacco Saudi/UAE","Alto"],
    ["Black Swan","N/A","N/D","N/D","Escluso per costruzione","â€”"],
],[4*cm,2*cm,3*cm,3.5*cm,6.5*cm,2.5*cm],hdr=["Scenario","Prob.","Brent","Equity","Trigger","Rischio"]))
story.append(Paragraph("Stima qualitativa esperta - non modello quantitativo formale.",SS))
story.append(Spacer(1,10))

story+=sec("INDICATORI DI RISCHIO","D")
def rs(val,ta,tr,hi=True):
    try:
        f=float(str(val).replace("%","").replace(",",".").replace("$","").replace("bp",""))
        if hi: return Paragraph("ALERT" if f>=tr else ("WARN" if f>=ta else "OK"),SR if f>=tr else (SA if f>=ta else SG))
        else:  return Paragraph("ALERT" if f<=tr else ("WARN" if f<=ta else "OK"),SR if f<=tr else (SA if f<=ta else SG))
    except: return Paragraph("N/D",SB)
story.append(tbl([
    ["VIX",str(vol.get("VIX","N/D")),"> 25","> 35",rs(vol.get("VIX",0),25,35)],
    ["Spread BTP-Bund",str(bonds.get("SPREAD_BTP_BUND","N/D")),"> 100bp","> 150bp",rs(bonds.get("SPREAD_BTP_BUND",0),100,150)],
    ["Brent",str(comm.get("BRENT","N/D")),"> $100","> $130",rs(comm.get("BRENT",0),100,130)],
    ["CDS Italia 5Y",str(bonds.get("CDS_ITALIA_5Y","N/D")),"> 35bp","> 60bp",rs(bonds.get("CDS_ITALIA_5Y",0),35,60)],
    ["EUR/USD",str(forex.get("EUR_USD","N/D")),"< 1.05","< 1.00",rs(forex.get("EUR_USD",1.10),1.05,1.00,False)],
    ["iTraxx HY","N/D","> 400bp","> 600bp",Paragraph("N/D - abbonamento",SB)],
],[4.5*cm,3.5*cm,3.5*cm,3.5*cm,5*cm],hdr=["Indicatore","Valore","Soglia AMBER","Soglia RED","Status"]))
story.append(Spacer(1,10))

story+=sec("AGENDA MACRO & FOMC","E")
story.append(tbl([
    ["Jun 16-17 *","FOMC","USA","SEP + Dot Plot","Fed.gov T1"],
    ["Jun 5","BCE","EU","Potenziale +25bp","ECB T1"],
    ["Jul 28-29","FOMC","USA","Standard meeting","Fed.gov T1"],
    ["Sep 15-16 *","FOMC","USA","SEP + Dot Plot","Fed.gov T1"],
    ["Oct 27-28","FOMC","USA","Standard meeting","Fed.gov T1"],
    ["Dec 8-9 *","FOMC","USA","SEP + Dot Plot","Fed.gov T1"],
],[3.5*cm,4*cm,2*cm,8*cm,4*cm],hdr=["Data","Evento","Area","Impatto atteso","Fonte"]))

for out in [PDF_PATH,PDF_OUT]:
    doc=SimpleDocTemplate(out,pagesize=landscape(A4),leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=2.0*cm,bottomMargin=1.5*cm)
    doc.build(story,onFirstPage=hdr_ftr,onLaterPages=hdr_ftr)
    print(f"OK WM PDF: {out}")

# XLSX
brd=Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))
def hs(c):
    c.font=Font(bold=True,color="FFFFFF",size=9); c.fill=PatternFill("solid",fgColor=NAVY_H)
    c.border=brd; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def ds(c,alt=False):
    c.font=Font(size=9); c.fill=PatternFill("solid",fgColor="F5F7FA") if alt else PatternFill()
    c.border=brd; c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
def aw(ws,mn=10,mx=40):
    for col in ws.columns:
        ml=mn
        for cell in col:
            try:
                if cell.value: ml=min(mx,max(mn,len(str(cell.value))+2))
            except: pass
        ws.column_dimensions[get_column_letter(col[0].column)].width=ml

wb=Workbook()
ws1=wb.active; ws1.title="Market Data"; ws1.sheet_properties.tabColor=NAVY_H
ws1["A1"]=f"FINANCIAL DAILY REPORT - Chiusura {DATE_CHIUSURA} | Report {DATE_REPORT}"
ws1["A1"].font=Font(bold=True,size=11,color=NAVY_H); ws1.merge_cells("A1:G1"); ws1.row_dimensions[1].height=20
for j,h in enumerate(["Asset","Categoria","Valore","Delta%","Fonte","Tier","Qualita"],1): hs(ws1.cell(row=2,column=j,value=h))
mdr=[
    ["FTSE MIB","Equity EU",v(idx,"FTSE_MIB"),v(idx,"FTSE_MIB_DELTA"),"MilanoFinanza","T2","Alta"],
    ["DAX","Equity EU",v(idx,"DAX"),v(idx,"DAX_DELTA"),"MilanoFinanza","T2","Alta"],
    ["S&P 500","Equity US",v(idx,"SP500"),v(idx,"SP500_DELTA"),"CNBC","T2","Alta"],
    ["Nasdaq","Equity US",v(idx,"NASDAQ"),v(idx,"NASDAQ_DELTA"),"CNBC","T2","Alta"],
    ["EUR/USD","FX",v(forex,"EUR_USD"),"N/D","ECB XML","T1","Alta"],
    ["EUR/JPY","FX",v(forex,"EUR_JPY"),"N/D","ECB XML","T1","Alta"],
    ["EUR/GBP","FX",v(forex,"EUR_GBP"),"N/D","ECB XML","T1","Alta"],
    ["BTP 10Y","Bond IT",v(bonds,"BTP_10Y"),"N/D","MilanoFinanza","T2","Alta"],
    ["Bund 10Y","Bond DE",v(bonds,"BUND_10Y"),"N/D","web_search","T2","Alta"],
    ["Spread","Spread",v(bonds,"SPREAD_BTP_BUND"),"N/D","calcolato","â€”","Alta"],
    ["UST 2Y","Bond US",v(bonds,"UST_2Y"),"N/D","web_search","T2","Alta"],
    ["UST 10Y","Bond US",v(bonds,"UST_10Y"),"N/D","web_search","T2","Alta"],
    ["UST 30Y","Bond US",v(bonds,"UST_30Y"),"N/D","web_search","T2","Alta"],
    ["VIX","Volatilita",v(vol,"VIX"),"N/D","CBOE","T1","Alta"],
    ["Brent","Commodity",v(comm,"BRENT"),v(comm,"BRENT_DELTA"),"CNBC","T2","Alta"],
    ["Gold","Commodity",v(comm,"GOLD"),v(comm,"GOLD_DELTA"),"CNBC","T2","Alta"],
    ["iTraxx HY","Credit","N/D","N/D","IHS Markit","T1","N/D"],
    ["BTC/USD","Crypto",v(crypto,"BTC_USD"),v(crypto,"BTC_DELTA"),"Composio","T2","Media"],
]
for i,row in enumerate(mdr,3):
    alt=i%2==0
    for j,val in enumerate(row,1):
        cell=ws1.cell(row=i,column=j,value=val); ds(cell,alt)
        if j==7 and val=="Alta": cell.font=Font(size=9,color=GREEN_H,bold=True)
        elif j==7 and "N/D" in str(val): cell.font=Font(size=9,color=AMBER_H)
aw(ws1)

ws2=wb.create_sheet("Raccomandazioni WM"); ws2.sheet_properties.tabColor=GOLD_H
ws2["A1"]=f"RACCOMANDAZIONI WM - {DATE_REPORT}"; ws2["A1"].font=Font(bold=True,size=11,color=NAVY_H); ws2.merge_cells("A1:E1")
for j,h in enumerate(["Asset Class","Posizionamento","Direzione","Rationale","Rischio"],1): hs(ws2.cell(row=2,column=j,value=h))
recs=[["Equity EU","Neutrale","->","Spread stabile; FTSE MIB resistenza 50k","Geopolitica"],
      ["Equity US","Sottopeso","<-","Valutazioni stretched; insider selling","Fed hawkish"],
      ["Fixed Income EU","Sovrappeso","->","BTP carry; BCE hawkish supporta","Duration"],
      ["Fixed Income US","Neutrale","->","UST 30Y >5%: opportunita selettiva","Deficit"],
      ["Commodities Oil","Neutrale","->","Hormuz: scenario B dominante","Geopolitica"],
      ["Gold","Sovrappeso","->","Safe haven + de-dollarizzazione","Liquidita"],
      ["Cash","Neutrale","->","EUR-STR ~2.9%","â€”"]]
for i,row in enumerate(recs,3):
    for j,val in enumerate(row,1):
        cell=ws2.cell(row=i,column=j,value=val); ds(cell,i%2==0)
        if j==2 and val=="Sovrappeso": cell.font=Font(size=9,color=GREEN_H,bold=True)
        elif j==2 and val=="Sottopeso": cell.font=Font(size=9,color=RED_H,bold=True)
aw(ws2)

ws3=wb.create_sheet("Risk Metrics"); ws3.sheet_properties.tabColor=RED_H
ws3["A1"]=f"RISK INDICATORS - {DATE_REPORT}"; ws3["A1"].font=Font(bold=True,size=11,color=NAVY_H); ws3.merge_cells("A1:E1")
for j,h in enumerate(["Metrica","Valore","Soglia AMBER","Soglia RED","Status"],1): hs(ws3.cell(row=2,column=j,value=h))
def rst(val,ta,tr,hi=True):
    try:
        f=float(str(val).replace("%","").replace(",",".").replace("$","").replace("bp",""))
        if hi: return "ALERT" if f>=tr else ("WARN" if f>=ta else "OK")
        else:  return "ALERT" if f<=tr else ("WARN" if f<=ta else "OK")
    except: return "N/D"
risk3=[
    ["VIX",str(vol.get("VIX","N/D")),"> 25","> 35",rst(vol.get("VIX",0),25,35)],
    ["Spread BTP-Bund",str(bonds.get("SPREAD_BTP_BUND","N/D")),"> 100bp","> 150bp",rst(bonds.get("SPREAD_BTP_BUND",0),100,150)],
    ["Brent",str(comm.get("BRENT","N/D")),"> $100","> $130",rst(comm.get("BRENT",0),100,130)],
    ["CDS Italia 5Y",str(bonds.get("CDS_ITALIA_5Y","N/D")),"> 35bp","> 60bp",rst(bonds.get("CDS_ITALIA_5Y",0),35,60)],
    ["EUR/USD",str(forex.get("EUR_USD","N/D")),"< 1.05","< 1.00",rst(forex.get("EUR_USD",1.10),1.05,1.00,False)],
    ["iTraxx HY","N/D","> 400bp","> 600bp","N/D - abbonamento"],
]
for i,row in enumerate(risk3,3):
    for j,val in enumerate(row,1):
        cell=ws3.cell(row=i,column=j,value=val); ds(cell,i%2==0)
        if j==5:
            if "ALERT" in str(val): cell.font=Font(size=9,color=RED_H,bold=True)
            elif "WARN" in str(val): cell.font=Font(size=9,color=AMBER_H,bold=True)
            elif "OK" in str(val): cell.font=Font(size=9,color=GREEN_H,bold=True)
aw(ws3)

for out in [XLSX_PATH,XLSX_OUT]: wb.save(out); print(f"OK XLSX: {out}")
print(f"\nWM Appendix + XLSX {DATE_REPORT} completati.")
